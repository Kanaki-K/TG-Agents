"""Строит таблицу аналитики постов канала из выгрузки.

Понимает два источника (берёт первый найденный, можно указать явно):
  1) data/result.json        — экспорт из Telegram Desktop (есть реакции, НЕТ просмотров)
  2) data/channel_posts.json — сбор через MTProto (collect.py): есть просмотры, репосты, комменты

Запуск:
    python -m connectors.telegram_export.build_table            # авто-выбор источника
    python -m connectors.telegram_export.build_table result     # из result.json
    python -m connectors.telegram_export.build_table mtproto    # из channel_posts.json

Результат: data/posts_analytics.xlsx (если есть openpyxl) и data/posts_analytics.csv
"""
from __future__ import annotations

import csv
import json
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data"
RESULT_JSON = DATA / "result.json"          # Telegram Desktop
MTPROTO_JSON = DATA / "channel_posts.json"  # collect.py
STATS_JSON = DATA / "channel_stats.json"    # collect_stats.py (админская сводка)
TOPICS_JSON = DATA / "post_topics.json"     # enrich_topics.py (заголовок/тема/суть)
SNAPS_JSON = DATA / "snapshots.json"        # snapshot.py (динамика метрик во времени)
OUT_XLSX = DATA / "posts_analytics.xlsx"
OUT_CSV = DATA / "posts_analytics.csv"

WEEKDAYS = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

# Перевод служебных меток Telegram → по-русски
LABELS = {
    "PM": "Личные сообщения", "Channels": "Каналы", "Followers": "Лента подписок",
    "Other": "Другое", "Groups": "Группы", "URL": "Ссылки", "Ads": "Реклама",
    "Search": "Поиск", "Shareable Chat Folders": "Папки-подборки",
    "Positive": "Позитивные", "Negative": "Негативные",
}


def _series_totals(graph: dict) -> list[tuple[str, float]]:
    """Сумма по каждой серии графика → [(метка, сумма)], по убыванию."""
    cols = graph.get("columns", [])
    names = graph.get("names", {})
    out = []
    for c in cols:
        if not c or c[0] == "x":
            continue
        label = LABELS.get(names.get(c[0], c[0]), names.get(c[0], c[0]))
        out.append((label, sum(v for v in c[1:] if isinstance(v, (int, float)))))
    out.sort(key=lambda x: -x[1])
    return out


def _flatten_text(text) -> str:
    """text в экспорте — строка ИЛИ список из строк и объектов {type,text,...}."""
    if isinstance(text, str):
        return text
    parts = []
    for chunk in text or []:
        if isinstance(chunk, str):
            parts.append(chunk)
        elif isinstance(chunk, dict):
            parts.append(chunk.get("text", ""))
    return "".join(parts)


def _content_type(m: dict) -> str:
    if "poll" in m:
        return "Опрос"
    if "giveaway_information" in m or "giveaway_results" in m:
        return "Розыгрыш"
    if m.get("sticker_emoji") or m.get("media_type") == "sticker":
        return "Стикер"
    mt = m.get("media_type")
    if mt in ("video_file", "video_message"):
        return "Видео"
    if mt == "voice_message":
        return "Голос"
    if mt == "animation":
        return "GIF"
    if "photo" in m:
        return "Фото"
    if "file" in m:
        return "Документ"
    return "Текст"


def _reactions_summary(reactions: list[dict]):
    """→ (всего, видов, топ-эмодзи, 'детально строкой')."""
    if not reactions:
        return 0, 0, "", ""
    pairs = []
    for r in reactions:
        emoji = r.get("emoji") or r.get("document_id") or "?"
        pairs.append((str(emoji), int(r.get("count", 0))))
    pairs.sort(key=lambda x: -x[1])
    total = sum(c for _, c in pairs)
    top = pairs[0][0] if pairs else ""
    detail = " ".join(f"{e}{c}" for e, c in pairs)
    return total, len(pairs), top, detail


def _rows_from_result(msgs: list[dict]) -> list[dict]:
    rows = []
    for m in msgs:
        if m.get("type") != "message":
            continue
        text = _flatten_text(m.get("text"))
        dt = datetime.fromisoformat(m["date"])
        total, kinds, top, detail = _reactions_summary(m.get("reactions", []))
        rows.append({
            "ID": m["id"],
            "Дата": dt.strftime("%Y-%m-%d"),
            "Время": dt.strftime("%H:%M"),
            "День недели": WEEKDAYS[dt.weekday()],
            "Час": dt.hour,
            "Тип контента": _content_type(m),
            "Репост из": m.get("forwarded_from", ""),
            "Длина (симв.)": len(text),
            "Слов": len(text.split()),
            "Реакций всего": total,
            "Видов реакций": kinds,
            "Топ-реакция": top,
            "Реакции (детально)": detail,
            "Просмотры": "",      # нет в Desktop-экспорте
            "Репосты": "",        # нет в Desktop-экспорте
            "Комментариев": "",   # нет в Desktop-экспорте
            "Редактирован": (m.get("edited") or "")[:10],
            "Превью": text[:80].replace("\n", " "),
            "Полный текст": text,
        })
    return rows


def _rows_from_mtproto(posts: list[dict]) -> list[dict]:
    rows = []
    for p in posts:
        text = p.get("text", "")
        dt = datetime.fromisoformat(p["date"]) if p.get("date") else None
        total, kinds, top, detail = _reactions_summary(p.get("reactions", []))
        views = p.get("views") or 0
        er = round(total / views * 100, 2) if views else ""
        rows.append({
            "ID": p["id"],
            "Дата": dt.strftime("%Y-%m-%d") if dt else "",
            "Время": dt.strftime("%H:%M") if dt else "",
            "День недели": WEEKDAYS[dt.weekday()] if dt else "",
            "Час": dt.hour if dt else "",
            "Тип контента": "Медиа" if p.get("has_media") else "Текст",
            "Репост из": "",
            "Длина (симв.)": len(text),
            "Слов": len(text.split()),
            "Реакций всего": total,
            "Видов реакций": kinds,
            "Топ-реакция": top,
            "Реакции (детально)": detail,
            "Просмотры": views,
            "Репосты": p.get("forwards") or 0,
            "Комментариев": p.get("comments") if p.get("comments") is not None else "",
            "ER % (реакции/просмотры)": er,
            "Редактирован": "",
            "Превью": text[:80].replace("\n", " "),
            "Полный текст": text,
        })
    return rows


def _apply_topics(rows: list[dict]) -> None:
    """Вставить колонки Заголовок/Тема (после «Тип контента») и Суть (перед текстом)."""
    if not TOPICS_JSON.exists():
        return
    topics = json.loads(TOPICS_JSON.read_text(encoding="utf-8"))
    for i, r in enumerate(rows):
        t = topics.get(str(r["ID"]), {})
        new = {}
        for k, v in r.items():
            if k == "Полный текст":
                new["Суть"] = t.get("summary", "")
            new[k] = v
            if k == "Тип контента":
                new["Заголовок"] = t.get("title", "")
                new["Тема"] = t.get("theme", "")
        rows[i] = new


def _write_csv(rows: list[dict]) -> None:
    cols = list(rows[0].keys())
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(rows)


def _write_xlsx(rows: list[dict]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        return False
    cols = list(rows[0].keys())
    wb = Workbook()
    ws = wb.active
    ws.title = "Посты"
    ws.append(cols)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in rows:
        ws.append([r[c] for c in cols])
    widths = {"Превью": 45, "Полный текст": 80, "Реакции (детально)": 22, "Репост из": 18,
              "Заголовок": 34, "Тема": 16, "Суть": 50}
    for i, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 13)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"

    if STATS_JSON.exists():
        _add_stats_sheets(wb)
    if SNAPS_JSON.exists():
        _add_dynamics_sheet(wb)

    wb.save(OUT_XLSX)
    return True


def _add_dynamics_sheet(wb) -> None:
    """Лист «Динамика 7 дней»: как росли метрики поста по дням после публикации."""
    from openpyxl.styles import Font, PatternFill

    snaps = json.loads(SNAPS_JSON.read_text(encoding="utf-8"))
    if not snaps:
        return
    topics = json.loads(TOPICS_JSON.read_text(encoding="utf-8")) if TOPICS_JSON.exists() else {}
    ws = wb.create_sheet("Динамика 7 дней")
    headers = ["Дата снимка", "Пост ID", "Заголовок", "Возраст (дн)",
               "Просмотры", "Реакции", "Комментарии", "Репосты"]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for s in sorted(snaps, key=lambda x: (x["post_id"], x.get("age_days", 0))):
        title = topics.get(str(s["post_id"]), {}).get("title", "")
        ws.append([s.get("date"), s["post_id"], title, s.get("age_days"),
                   s.get("views"), s.get("reactions"), s.get("comments"), s.get("forwards")])
    ws.freeze_panes = "A2"
    ws.column_dimensions["C"].width = 34
    for col in "ABDEFGH":
        ws.column_dimensions[col].width = 13


def _add_stats_sheets(wb) -> None:
    """Добавить листы со сводной статистикой канала из channel_stats.json."""
    from datetime import datetime, timezone
    from openpyxl.styles import Font, PatternFill

    stats = json.loads(STATS_JSON.read_text(encoding="utf-8"))
    head_fill = PatternFill("solid", fgColor="1F4E78")

    def new_sheet(title, headers):
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(1, c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = head_fill
        ws.column_dimensions["A"].width = 28
        for col in "BCD":
            ws.column_dimensions[col].width = 16
        return ws

    # 1) Сводка канала (заголовочные цифры)
    h = stats.get("headline", {})
    titles = {
        "followers": "Подписчики", "views_per_post": "Просмотров на пост",
        "shares_per_post": "Репостов на пост", "reactions_per_post": "Реакций на пост",
        "views_per_story": "Просмотров на историю",
    }
    ws = new_sheet("Сводка канала", ["Метрика", "Сейчас", "Было", "Изм. %"])
    p = stats.get("period", {})
    ws.append([f"Период: {(p.get('from') or '')[:10]} → {(p.get('to') or '')[:10]}", "", "", ""])
    for key, name in titles.items():
        v = h.get(key)
        if not v:
            continue
        ws.append([name, v.get("current"), v.get("previous"), v.get("change_pct")])

    # 2) Листы-разбивки (сумма по сериям)
    breakdowns = [
        ("Просмотры по источникам", "Просмотры по источникам", "Просмотры"),
        ("Новые подписчики по источникам", "Новые подписчики", "Подписки"),
        ("Языки аудитории", "Языки аудитории", "Просмотры"),
        ("Реакции по эмоциям", "Реакции по эмоциям", "Кол-во"),
    ]
    graphs = stats.get("graphs", {})
    for gname, title, valcol in breakdowns:
        g = graphs.get(gname)
        if not g or "columns" not in g:
            continue
        ws = new_sheet(title, ["Категория", valcol])
        for label, total in _series_totals(g):
            ws.append([label, int(total)])

    # 3) Активность по часам (последний период)
    g = graphs.get("Активность по часам")
    if g and g.get("columns"):
        cols = g["columns"]
        ws = new_sheet("Активность по часам", ["Час", "Просмотры (посл. период)"])
        x, y = cols[0][1:], (cols[1][1:] if len(cols) > 1 else [])
        for hour, val in zip(x, y):
            ws.append([f"{hour:02d}:00", val])

    # 4) Рост аудитории (динамика по дням)
    g = graphs.get("Рост аудитории")
    if g and g.get("columns"):
        cols = g["columns"]
        ws = new_sheet("Рост аудитории", ["Дата", "Подписчиков"])
        x = cols[0][1:]
        y = cols[1][1:] if len(cols) > 1 else []
        for ts, val in zip(x, y):
            d = datetime.fromtimestamp(ts / 1000, tz=timezone.utc).strftime("%Y-%m-%d")
            ws.append([d, val])


def main() -> None:
    src = sys.argv[1] if len(sys.argv) > 1 else "auto"
    if src in ("mtproto", "auto") and MTPROTO_JSON.exists():
        posts = json.loads(MTPROTO_JSON.read_text(encoding="utf-8"))
        rows = _rows_from_mtproto(posts)
        used = MTPROTO_JSON.name
    elif RESULT_JSON.exists():
        data = json.loads(RESULT_JSON.read_text(encoding="utf-8"))
        rows = _rows_from_result(data["messages"])
        used = RESULT_JSON.name
    else:
        raise SystemExit("Нет источника: положи data/result.json или собери data/channel_posts.json")

    rows.sort(key=lambda r: (r["Дата"], r["Время"]))
    _apply_topics(rows)
    _write_csv(rows)
    xlsx_ok = _write_xlsx(rows)

    total_react = sum(r["Реакций всего"] for r in rows)
    print(f"Источник: {used}")
    print(f"Постов в таблице: {len(rows)}")
    print(f"Суммарно реакций: {total_react}")
    print(f"CSV  → {OUT_CSV}")
    print(f"XLSX → {OUT_XLSX}" if xlsx_ok else "XLSX пропущен (нет openpyxl) — открой CSV в Excel")


if __name__ == "__main__":
    main()

"""Строит таблицу аналитики постов Threads — аналог telegram_export/build_table.py.

Источники (готовят предыдущие шаги):
  data/threads_posts.json  — посты + метрики (collect.py)
  data/threads_topics.json — заголовок/тема/суть (enrich_topics.py)
  data/threads_stats.json  — сводка аккаунта + демография (collect.py)

Результат: data/threads_analytics.xlsx (если есть openpyxl) и data/threads_analytics.csv —
той же структуры, что ТГ-таблица, чтобы аналитик делал такие же глубокие выводы.

Запуск:
    python -m connectors.threads.build_table
"""
from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path

from connectors.threads import scoring

ROOT = Path(__file__).resolve().parents[2]
DATA = ROOT / "data"
POSTS_JSON = DATA / "threads_posts.json"
TOPICS_JSON = DATA / "threads_topics.json"
STATS_JSON = DATA / "threads_stats.json"
OUT_XLSX = DATA / "threads_analytics.xlsx"
OUT_CSV = DATA / "threads_analytics.csv"

WEEKDAYS = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]

# Тип медиа Threads → по-русски
MEDIA = {
    "TEXT_POST": "Текст", "TEXT": "Текст", "IMAGE": "Фото", "VIDEO": "Видео",
    "CAROUSEL_ALBUM": "Карусель", "AUDIO": "Аудио", "REPOST_FACADE": "Репост",
}

DEMO_TITLES = {"country": "Подписчики по странам", "city": "Подписчики по городам",
               "age": "Подписчики по возрасту", "gender": "Подписчики по полу"}


def _content_type(p: dict) -> str:
    t = MEDIA.get(p.get("media_type", ""), p.get("media_type") or "Текст")
    return f"{t} (цитата)" if p.get("is_quote") else t


def _rows(posts: list[dict]) -> list[dict]:
    rows = []
    for p in posts:
        text = p.get("text", "") or ""
        raw = p.get("date", "")
        dt = None
        if raw:
            try:
                dt = datetime.fromisoformat(raw)
            except ValueError:
                dt = None
        rows.append({
            "ID": p.get("id", ""),
            "Дата": dt.strftime("%Y-%m-%d") if dt else "",
            "Время": dt.strftime("%H:%M") if dt else "",
            "День недели": WEEKDAYS[dt.weekday()] if dt else "",
            "Час": dt.hour if dt else "",
            "Тип контента": _content_type(p),
            "С медиа": "да" if p.get("has_media") else "",
            "Ссылка в тексте": "да" if p.get("has_link") else "",
            "Хештег": "да" if p.get("has_hashtag") else "",
            "Длина (симв.)": len(text),
            "Слов": len(text.split()),
            "Просмотры": p.get("views", 0),
            "Лайки": p.get("likes", 0),
            "Ответы (всего)": p.get("replies", 0),
            "Комментов людей": p.get("people_replies", 0),
            "Людей в комментах": p.get("people_count", 0),
            "Свои ответы": p.get("self_replies", 0),
            "Реплаев на человека": p.get("replies_per_person", 0),
            "Людей/1k показов": p.get("people_per_k") if p.get("people_per_k") is not None else "",
            "Репосты": p.get("reposts", 0),
            "Цитаты": p.get("quotes", 0),
            "Пересылки": p.get("shares", 0),
            "Распространений": p.get("amplification", 0),
            "Вовлечённость": p.get("engagement", 0),
            "ER % (вовл./просмотры)": p.get("er") if p.get("er") is not None else "",
            "Возраст (дн)": p.get("age_days") if p.get("age_days") is not None else "",
            "Качество (0-100)": p.get("quality") if p.get("quality") is not None else "",
            "Охват %": p.get("reach_pct") if p.get("reach_pct") is not None else "",
            "Тип отклика": p.get("tier_ru", ""),
            "Ссылка": p.get("permalink", ""),
            "Превью": text[:80].replace("\n", " "),
            "Полный текст": text,
        })
    return rows


def _apply_topics(rows: list[dict]) -> None:
    """Вставить Заголовок/Тема (после «Тип контента») и Суть (перед полным текстом)."""
    if not TOPICS_JSON.exists():
        return
    topics = json.loads(TOPICS_JSON.read_text(encoding="utf-8"))
    for i, r in enumerate(rows):
        t = topics.get(str(r["ID"]), {})
        new = {}
        for k, v in r.items():
            if k == "Полный текст":
                new["Суть"] = t.get("summary", "")
            new[k] = v
            if k == "Тип контента":
                new["Заголовок"] = t.get("title", "")
                new["Тема"] = t.get("theme", "")
        rows[i] = new


def _write_csv(rows: list[dict]) -> None:
    cols = list(rows[0].keys())
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerows(rows)


def _write_xlsx(rows: list[dict]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError:
        return False
    cols = list(rows[0].keys())
    wb = Workbook()
    ws = wb.active
    ws.title = "Посты"
    ws.append(cols)
    head_fill = PatternFill("solid", fgColor="1F4E78")
    for c in range(1, len(cols) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for r in rows:
        ws.append([r[c] for c in cols])
    widths = {"Превью": 45, "Полный текст": 80, "Ссылка": 30,
              "Заголовок": 34, "Тема": 16, "Суть": 50}
    for i, name in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"

    if STATS_JSON.exists():
        _add_stats_sheets(wb)

    wb.save(OUT_XLSX)
    return True


def _add_stats_sheets(wb) -> None:
    from openpyxl.styles import Font, PatternFill

    stats = json.loads(STATS_JSON.read_text(encoding="utf-8"))
    head_fill = PatternFill("solid", fgColor="1F4E78")

    def new_sheet(title, headers):
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(1, c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = head_fill
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 16
        return ws

    # Сводка аккаунта за период
    acc = stats.get("account") or {}
    p = stats.get("period", {})

    def _fmt(ts):
        try:
            return datetime.fromtimestamp(int(ts)).strftime("%Y-%m-%d")
        except Exception:  # noqa: BLE001
            return ""

    ws = new_sheet("Сводка аккаунта", ["Метрика", "Значение"])
    ws.append([f"Период: {_fmt(p.get('from'))} → {_fmt(p.get('to'))}", ""])
    names = {"views": "Просмотры", "likes": "Лайки", "replies": "Ответы",
             "reposts": "Репосты", "quotes": "Цитаты", "followers_count": "Подписчики"}
    if acc.get("error"):
        ws.append(["⚠ сводка недоступна", acc["error"]])
    else:
        for key, name in names.items():
            if acc.get(key) is not None:
                ws.append([name, acc[key]])

    # Демография подписчиков (если пришла — нужны ≥100 подписчиков + привязка к Instagram)
    for key, breakdown in (stats.get("demographics") or {}).items():
        if not isinstance(breakdown, dict) or breakdown.get("error") or not breakdown:
            continue
        ws = new_sheet(DEMO_TITLES.get(key, key), ["Категория", "Подписчиков"])
        for label, val in sorted(breakdown.items(), key=lambda x: -(x[1] or 0)):
            ws.append([label, val])


def main() -> None:
    if not POSTS_JSON.exists():
        raise SystemExit("Нет data/threads_posts.json — сначала: python -m connectors.threads.collect")
    posts = json.loads(POSTS_JSON.read_text(encoding="utf-8"))
    if not posts:
        raise SystemExit("В data/threads_posts.json нет постов.")
    scoring.enrich(posts)   # производные метрики, тир, Сила — единая логика с отчётом
    rows = _rows(posts)
    rows.sort(key=lambda r: (r["Дата"], r["Время"]))
    _apply_topics(rows)
    _write_csv(rows)
    xlsx_ok = _write_xlsx(rows)

    total_views = sum(r["Просмотры"] for r in rows)
    total_eng = sum(r["Вовлечённость"] for r in rows)
    print(f"Постов в таблице: {len(rows)}")
    print(f"Суммарно просмотров: {total_views}, вовлечённости: {total_eng}")
    print(f"CSV  → {OUT_CSV}")
    print(f"XLSX → {OUT_XLSX}" if xlsx_ok else "XLSX пропущен (нет openpyxl) — открой CSV в Excel")


if __name__ == "__main__":
    main()
